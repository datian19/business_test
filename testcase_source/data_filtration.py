# coding:utf-8
import time

import openpyxl
from openpyxl import load_workbook

'''
路测数据处理
'''


class Data_Filtration:
    def __init__(self):
        self.title_row = None
        self.statedata_callback = None
        self.return_data = []

    # 路测数据
    def run_data_edit(self, inFilePaths, outFilePath):
        self.statedata_callback("路测数据处理开始")
        rowItems = []
        for infile in inFilePaths:
            dictRead = self.read_excel(infile)
            if dictRead.get("retCode") == 1:
                self.statedata_callback(f'路测数据[{infile}]处理失败')
                return 1
            retData = dictRead.get("retData")
            rowItems.extend(retData)

        validItems = self.edit_data(rowItems)
        dictWrite = self.write_excel(outFilePath, validItems)
        if dictWrite.get("retCode") == 1:
            return 1
        self.statedata_callback("路测数据处理结束")
        return 0

    # 读取excel文件
    def read_excel(self, inFilePath):
        # 加载excel文件
        self.statedata_callback(f'读取路测文件[{inFilePath}]......')
        try:
            wbook = load_workbook(inFilePath)
            # 获取sheet列表
            sheetNames = wbook.get_sheet_names()
            # 获取第一个sheet的数据
            sheetData = wbook.get_sheet_by_name(sheetNames[0])
            datas = list(sheetData.rows)
            self.title_row = [cell.value for cell in datas[0]]
        except:
            self.statedata_callback("文件读取失败......")
            return {"retCode": 1, "retData": None}
        # 获取每一行的数据
        rowItems = []
        for row in datas[1:]:
            item = [cell.value for cell in row]
            rowItems.append(item)
        return {"retCode": 0, "retData": rowItems}

    # 编辑数据
    def edit_data(self, rowItems):
        self.statedata_callback("路测数据编辑开始......")
        dataCnt = len(rowItems)  # 数据总件数
        cgiCnt = 0  # CGI是空的件数
        lonCnt = 0  # 经度(longitude)是空的件数
        latCnt = 0  # 纬度(Latitude)是空的件数
        rsrpCnt = 0  # RSRP是空的件数
        validItems = []
        for item in rowItems:
            if item[5] is None or str(item[5]).strip() == "":  # CGI是空的判断
                cgiCnt += 1
            elif item[6] is None or str(item[6]).strip() == "":  # 经度(longitude)是空的判断
                lonCnt += 1
            elif item[7] is None or str(item[7]).strip() == "":  # 纬度(Latitude)是空的件数
                latCnt += 1
            elif item[9] is None or str(item[9]).strip() == "":  # RSRP是空的件数
                rsrpCnt += 1
            else:
                validItems.append(item)

        # 去除重复数据（PCI,经度，纬度重复的数据计算RSRP的平均值）
        data_dict = {}
        for item in validItems:
            # PCI,经度，纬度编辑组合key
            itemKey = "{0}{1}{2}".format(item[5], item[6], item[7])
            if not itemKey in data_dict:
                data_dict[itemKey] = [item]
            else:
                data_dict[itemKey].append(item)
        retDatList = []
        repetitionscnt = 0
        repetitionList = []
        for key, value in data_dict.items():
            if len(value) == 0:
                continue
            if len(value) > 1:
                rsrp = 0.0
                itemcnt = 0
                for item in value:
                    itemcnt += 1
                    rsrp += float(item[9])
                itemdata = value[0]
                itemdata[9] = round(rsrp / itemcnt, 3)
                retDatList.append(itemdata)
                repetitionscnt += (itemcnt - 1)
                repetitionList.append([itemdata[5], itemdata[6], itemdata[7], str(itemcnt)])
            else:
                retDatList.append(value[0])

        self.return_data = [dataCnt, cgiCnt, lonCnt, latCnt, rsrpCnt, repetitionscnt, len(retDatList), repetitionList]
        # print(len(self.return_data[7]))
        return retDatList

    # 写入excel文件
    def write_excel(self, outFilePath, rowItems):
        self.statedata_callback(f'路测文件[{outFilePath}]输出......')
        # 调用openpyxl的Workbook()函数以创建一个新的空白Workbook对象
        try:
            wbook = openpyxl.Workbook()
            ws_first = wbook.worksheets[0]
            ws_first.append(self.title_row)
            for item in rowItems:
                ws_first.append(item)
            time.sleep(2)
            wbook.save(outFilePath)
            return {"retCode": 0, "retData": None}
        except:
            self.statedata_callback("路测文件输出失败，请确认文件是否被打开！")
            return {"retCode": 1, "retData": None}

    def set_road_state(self, statedataFunc):
        self.statedata_callback = statedataFunc
        return
